import csv
import io
from datetime import date, datetime, time

import pytest
import pytest_asyncio
from openpyxl import load_workbook
from sqlalchemy import text

from app.auth import create_access_token, hash_password
from app.models import User


async def _create_user(db_session, email="user@example.com", password="Sup3rSecret!"):
    user = User(email=email, password_hash=hash_password(password), full_name="Test User")
    db_session.add(user)
    await db_session.commit()
    await db_session.refresh(user)
    return user


async def _auth_headers(client, db_session):
    await _create_user(db_session)
    response = await client.post(
        "/api/auth/login", json={"email": "user@example.com", "password": "Sup3rSecret!"}
    )
    token = response.json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


INSERT_PERSONA_SQL = text(
    """
    INSERT INTO axis.personas (identificacion, tipo_identificacion, nombre)
    VALUES (:identificacion, :tipo_identificacion, :nombre)
    ON CONFLICT (identificacion) DO NOTHING
    """
)

INSERT_SQL = text(
    """
    INSERT INTO axis.axis_libretines
        (registro, hora_generacion, codigo_libretin, prefijo_boleta, rango_inicio_boleta,
         rango_fin_boleta, cantidad_boletas, longitud_boleta, estado, codigo_tramite,
         codigo_usuario_creacion, codigo_tramite_asignacion, codigo_usuario_asignacion,
         codigo_usuario_inactiva, observacion, codigo_agente, identificacion_agente, agente,
         codigo_distrito, descripcion_distrito, codigo_oficina, descripcion_oficina,
         codigo_provincia, descripcion_provincia, codigo_localidad, descripcion_localidad, tipo,
         origen_tramite, motivo_baja, disponibles, utilizadas, desactivadas, fecha_generacion,
         fecha_registro, fecha_asignacion, fecha_inactivacion, codigo_localidad_catalogo_item_id,
         codigo_provincia_catalogo_item_id, estado_catalogo_item_id, tipo_catalogo_item_id)
    VALUES
        (:registro, :hora_generacion, :codigo_libretin, :prefijo_boleta, :rango_inicio_boleta,
         :rango_fin_boleta, :cantidad_boletas, :longitud_boleta, :estado, :codigo_tramite,
         :codigo_usuario_creacion, :codigo_tramite_asignacion, :codigo_usuario_asignacion,
         :codigo_usuario_inactiva, :observacion, :codigo_agente, :identificacion_agente, :agente,
         :codigo_distrito, :descripcion_distrito, :codigo_oficina, :descripcion_oficina,
         :codigo_provincia, :descripcion_provincia, :codigo_localidad, :descripcion_localidad, :tipo,
         :origen_tramite, :motivo_baja, :disponibles, :utilizadas, :desactivadas, :fecha_generacion,
         :fecha_registro, :fecha_asignacion, :fecha_inactivacion, :codigo_localidad_catalogo_item_id,
         :codigo_provincia_catalogo_item_id, :estado_catalogo_item_id, :tipo_catalogo_item_id)
    RETURNING id
    """
)


def _row(registro, fecha_registro, identificacion_agente="TEST-LIB-CED-0001", **overrides):
    base = {
        "registro": registro,
        "hora_generacion": time(7, 30, 0),
        "codigo_libretin": f"LIB-{registro}",
        "prefijo_boleta": "A",
        "rango_inicio_boleta": "000001",
        "rango_fin_boleta": "000100",
        "cantidad_boletas": "100",
        "longitud_boleta": "6",
        "estado": "ACTIVO",
        "codigo_tramite": "TRA-001",
        "codigo_usuario_creacion": "USR-001",
        "codigo_tramite_asignacion": "TRA-002",
        "codigo_usuario_asignacion": "USR-002",
        "codigo_usuario_inactiva": None,
        "observacion": "Observación de prueba",
        "codigo_agente": "AGT-001",
        "identificacion_agente": identificacion_agente,
        "agente": "Agente de Prueba",
        "codigo_distrito": "D-01",
        "descripcion_distrito": "Distrito Centro",
        "codigo_oficina": "OF-01",
        "descripcion_oficina": "Oficina Centro",
        "codigo_provincia": "LOJ",
        "descripcion_provincia": "Loja",
        "codigo_localidad": "LOJ-01",
        "descripcion_localidad": "Loja",
        "tipo": "NORMAL",
        "origen_tramite": "MANUAL",
        "motivo_baja": None,
        "disponibles": "50",
        "utilizadas": "40",
        "desactivadas": "10",
        "fecha_generacion": fecha_registro.date(),
        "fecha_registro": fecha_registro,
        "fecha_asignacion": None,
        "fecha_inactivacion": None,
        "codigo_localidad_catalogo_item_id": 8,
        "codigo_provincia_catalogo_item_id": 4,
        "estado_catalogo_item_id": 2,
        "tipo_catalogo_item_id": 1,
    }
    base.update(overrides)
    return base


async def _seed_personas(db_session, identificaciones):
    for ident in identificaciones:
        await db_session.execute(
            INSERT_PERSONA_SQL,
            {"identificacion": ident, "tipo_identificacion": "CED", "nombre": f"Persona {ident}"},
        )
    await db_session.commit()


async def _seed_libretines(db_session, rows):
    personas = {row["identificacion_agente"] for row in rows}
    await _seed_personas(db_session, personas)
    ids = []
    for row in rows:
        result = await db_session.execute(INSERT_SQL, row)
        ids.append(result.scalar_one())
    await db_session.commit()
    return ids


@pytest_asyncio.fixture(autouse=True)
async def _cleanup_libretines(db_session):
    yield
    await db_session.execute(text("DELETE FROM axis.axis_libretines WHERE registro LIKE 'TEST-LIB-%'"))
    await db_session.execute(text("DELETE FROM axis.personas WHERE identificacion LIKE 'TEST-LIB-%'"))
    await db_session.commit()


@pytest.mark.asyncio
async def test_list_returns_items_within_range(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_libretines(
        db_session,
        [
            _row("TEST-LIB-101", datetime(2031, 6, 5, 9, 0, 0), identificacion_agente="TEST-LIB-CED-0001"),
            _row("TEST-LIB-102", datetime(2031, 6, 15, 9, 0, 0), identificacion_agente="TEST-LIB-CED-0002"),
            _row("TEST-LIB-103", datetime(2031, 6, 25, 9, 0, 0), identificacion_agente="TEST-LIB-CED-0003"),
        ],
    )

    response = await client.get(
        "/api/reportes/libretines",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 3
    assert body["page"] == 1
    assert body["page_size"] == 50
    registros = [item["registro"] for item in body["items"]]
    assert registros == ["TEST-LIB-103", "TEST-LIB-102", "TEST-LIB-101"]
    first = body["items"][0]
    assert first["agente"] == "Agente de Prueba"
    assert first["identificacion_agente"] == "TEST-LIB-CED-0003"
    assert first["disponibles"] == "50"
    assert first["estado_catalogo_item_id"] == 2


@pytest.mark.asyncio
async def test_list_allows_range_crossing_month(client, db_session):
    headers = await _auth_headers(client, db_session)

    response = await client.get(
        "/api/reportes/libretines",
        params={"fecha_desde": "2031-06-15", "fecha_hasta": "2031-07-05"},
        headers=headers,
    )

    assert response.status_code == 200


@pytest.mark.asyncio
async def test_list_rejects_desde_after_hasta(client, db_session):
    headers = await _auth_headers(client, db_session)

    response = await client.get(
        "/api/reportes/libretines",
        params={"fecha_desde": "2031-06-20", "fecha_hasta": "2031-06-10"},
        headers=headers,
    )

    assert response.status_code == 400


@pytest.mark.asyncio
async def test_list_excludes_soft_deleted_rows(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_libretines(
        db_session,
        [
            _row("TEST-LIB-201", datetime(2031, 6, 5, 9, 0, 0), identificacion_agente="TEST-LIB-CED-0004"),
            _row("TEST-LIB-202", datetime(2031, 6, 6, 9, 0, 0), identificacion_agente="TEST-LIB-CED-0005"),
        ],
    )
    await db_session.execute(
        text("UPDATE axis.axis_libretines SET deleted_at = now() WHERE registro = 'TEST-LIB-202'")
    )
    await db_session.commit()

    response = await client.get(
        "/api/reportes/libretines",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 1
    assert body["items"][0]["registro"] == "TEST-LIB-201"


@pytest.mark.asyncio
async def test_list_includes_rows_at_the_end_of_the_last_day(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_libretines(
        db_session,
        [
            _row("TEST-LIB-601", datetime(2031, 6, 30, 23, 59, 59), identificacion_agente="TEST-LIB-CED-0006"),
            _row("TEST-LIB-602", datetime(2031, 7, 1, 0, 0, 0), identificacion_agente="TEST-LIB-CED-0007"),
        ],
    )

    response = await client.get(
        "/api/reportes/libretines",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    registros = [item["registro"] for item in body["items"]]
    assert "TEST-LIB-601" in registros
    assert "TEST-LIB-602" not in registros


@pytest.mark.asyncio
async def test_list_pagination_page_two_offset(client, db_session):
    headers = await _auth_headers(client, db_session)
    base = datetime(2031, 5, 1, 8, 0, 0)
    rows = [_row(f"TEST-LIB-p-{i:03d}", base.replace(day=1 + i % 28)) for i in range(55)]
    await _seed_libretines(db_session, rows)

    response = await client.get(
        "/api/reportes/libretines",
        params={"fecha_desde": "2031-05-01", "fecha_hasta": "2031-06-25", "page": 2},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 55
    assert body["page"] == 2
    assert len(body["items"]) == 5


@pytest.mark.asyncio
async def test_list_out_of_range_page_returns_empty(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_libretines(db_session, [_row("TEST-LIB-301", datetime(2031, 6, 5, 9, 0, 0))])

    response = await client.get(
        "/api/reportes/libretines",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30", "page": 5},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["items"] == []
    assert body["total"] == 1


@pytest.mark.asyncio
async def test_list_without_token_returns_401(client, db_session):
    response = await client.get(
        "/api/reportes/libretines",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
    )
    assert response.status_code == 401


@pytest.mark.asyncio
async def test_list_blocked_when_must_change_password_is_true(client, db_session):
    user = User(
        email="pendiente@example.com",
        password_hash=hash_password("Sup3rSecret!"),
        full_name="Usuario Pendiente",
        must_change_password=True,
    )
    db_session.add(user)
    await db_session.commit()
    await db_session.refresh(user)
    token = create_access_token(user_id=user.id, email=user.email)

    response = await client.get(
        "/api/reportes/libretines",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 403
    assert response.json()["detail"]["code"] == "password_change_required"


EXPECTED_HEADERS = [
    "Registro",
    "Hora de Generación del Registro",
    "Código Libretin",
    "Prefijo Boleta",
    "Rango Inicio Boleta",
    "Rango Fin Boleta",
    "Cantidad Boletas",
    "Longitud Boleta",
    "Estado",
    "Código de Trámite",
    "Código de Usuario Creación",
    "Código de Trámite Asignación",
    "Código de Usuario Asignación",
    "Código de Usuario Inactiva",
    "Observación",
    "Código Agente",
    "Identificación Agente",
    "Agente",
    "Código Distrito",
    "Descripción Distrito",
    "Código Oficina",
    "Descripción Oficina",
    "Código Provincia",
    "Descripción Provincia",
    "Código Localidad",
    "Descripción Localidad",
    "Tipo",
    "Origen Trámite",
    "Motivo Baja",
    "Disponibles",
    "Utilizadas",
    "Desactivadas",
    "Fecha de Generación del Registro",
    "Fecha de Registro",
    "Fecha Asignación",
    "Fecha Inactivación",
    "ID de Catálogo (Localidad)",
    "ID de Catálogo (Provincia)",
    "ID de Catálogo (Estado)",
    "ID de Catálogo (Tipo)",
]


@pytest.mark.asyncio
async def test_export_csv_returns_all_matching_rows(client, db_session):
    headers = await _auth_headers(client, db_session)
    rows = [
        _row(f"TEST-LIB-e-{i:03d}", datetime(2031, 8, 1 + i, 9, 0, 0))
        for i in range(28)
    ]
    await _seed_libretines(db_session, rows)

    response = await client.get(
        "/api/reportes/libretines/export",
        params={"fecha_desde": "2031-08-01", "fecha_hasta": "2031-08-31", "formato": "csv"},
        headers=headers,
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    assert "libretines_2031-08-01_2031-08-31.csv" in response.headers["content-disposition"]

    text_content = response.content.decode("utf-8-sig")
    lines = [line for line in text_content.splitlines() if line]
    reader = csv.reader(lines)
    parsed_rows = list(reader)
    assert parsed_rows[0] == EXPECTED_HEADERS
    assert len(parsed_rows[0]) == 40
    assert len(lines) - 1 == 28

    data_row = parsed_rows[1]
    assert len(data_row) == 40
    assert data_row[0].startswith("TEST-LIB-e-")


@pytest.mark.asyncio
async def test_export_xlsx_returns_all_matching_rows(client, db_session):
    headers = await _auth_headers(client, db_session)
    rows = [
        _row(f"TEST-LIB-x-{i:03d}", datetime(2031, 9, 1 + i, 9, 0, 0))
        for i in range(29)
    ]
    await _seed_libretines(db_session, rows)

    response = await client.get(
        "/api/reportes/libretines/export",
        params={"fecha_desde": "2031-09-01", "fecha_hasta": "2031-09-30", "formato": "xlsx"},
        headers=headers,
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "libretines_2031-09-01_2031-09-30.xlsx" in response.headers["content-disposition"]

    workbook = load_workbook(io.BytesIO(response.content))
    sheet = workbook.active
    header_row = [sheet.cell(row=1, column=col).value for col in range(1, 41)]
    assert header_row == EXPECTED_HEADERS
    assert sheet.max_row - 1 == 29

    data_row = [sheet.cell(row=2, column=col).value for col in range(1, 41)]
    assert data_row[0].startswith("TEST-LIB-x-")


@pytest.mark.asyncio
async def test_export_excludes_soft_deleted_rows(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_libretines(
        db_session,
        [
            _row("TEST-LIB-501", datetime(2031, 6, 5, 9, 0, 0), identificacion_agente="TEST-LIB-CED-0008"),
            _row("TEST-LIB-502", datetime(2031, 6, 6, 9, 0, 0), identificacion_agente="TEST-LIB-CED-0009"),
        ],
    )
    await db_session.execute(
        text("UPDATE axis.axis_libretines SET deleted_at = now() WHERE registro = 'TEST-LIB-502'")
    )
    await db_session.commit()

    response = await client.get(
        "/api/reportes/libretines/export",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30", "formato": "csv"},
        headers=headers,
    )

    assert response.status_code == 200
    text_content = response.content.decode("utf-8-sig")
    lines = [line for line in text_content.splitlines() if line]
    assert len(lines) - 1 == 1


@pytest.mark.asyncio
async def test_export_without_token_returns_401(client, db_session):
    response = await client.get(
        "/api/reportes/libretines/export",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30", "formato": "csv"},
    )
    assert response.status_code == 401


@pytest.mark.asyncio
async def test_list_truncates_datetime_columns_to_date_only(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_libretines(
        db_session,
        [
            _row(
                "TEST-LIB-TRUNC-001",
                datetime(2031, 6, 5, 14, 35, 0),
                fecha_asignacion=datetime(2031, 6, 6, 10, 0, 0),
                fecha_inactivacion=datetime(2031, 6, 10, 9, 0, 0),
            )
        ],
    )

    response = await client.get(
        "/api/reportes/libretines",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
        headers=headers,
    )

    assert response.status_code == 200
    first = response.json()["items"][0]
    assert first["fecha_generacion"] == "2031-06-05"
    assert first["fecha_registro"] == "2031-06-05"
    assert first["fecha_asignacion"] == "2031-06-06"
    assert first["fecha_inactivacion"] == "2031-06-10"


@pytest.mark.asyncio
async def test_export_truncates_datetime_columns_to_date_only(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_libretines(
        db_session,
        [
            _row(
                "TEST-LIB-TRUNC-002",
                datetime(2031, 6, 6, 14, 35, 0),
                fecha_asignacion=datetime(2031, 6, 7, 10, 0, 0),
                fecha_inactivacion=datetime(2031, 6, 11, 9, 0, 0),
            )
        ],
    )

    response = await client.get(
        "/api/reportes/libretines/export",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30", "formato": "csv"},
        headers=headers,
    )

    assert response.status_code == 200
    text_content = response.content.decode("utf-8-sig")
    lines = [line for line in text_content.splitlines() if line]
    reader = csv.reader(lines)
    parsed_rows = list(reader)
    data_row = parsed_rows[-1]
    assert data_row[32] == "2031-06-06"
    assert data_row[33] == "2031-06-06"
    assert data_row[34] == "2031-06-07"
    assert data_row[35] == "2031-06-11"
