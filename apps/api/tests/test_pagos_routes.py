import csv
import io
from datetime import date, datetime, time, timedelta

import pytest
import pytest_asyncio
from decimal import Decimal
from openpyxl import load_workbook
from sqlalchemy import text

from app.auth import create_access_token, hash_password
from app.models import User


async def _create_user(db_session, email="user@example.com", password="Sup3rSecret!"):
    user = User(email=email, password_hash=hash_password(password), full_name="Test User")
    db_session.add(user)
    await db_session.commit()
    await db_session.refresh(user)
    return user


async def _auth_headers(client, db_session):
    await _create_user(db_session)
    response = await client.post(
        "/api/auth/login", json={"email": "user@example.com", "password": "Sup3rSecret!"}
    )
    token = response.json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


INSERT_SQL = text(
    """
    INSERT INTO axis.axis_pagos
        (registro, hora_generacion, tipo_recaudador, recaudador, comprobante_pago_interno,
         comprobante_pago_recaudador, tipo_servicio, tipo_documento, numero_documento,
         fecha_generacion, fecha_operacion, fecha_transaccion, monto_recaudado,
         monto_cuenta_1, monto_cuenta_2)
    VALUES
        (:registro, :hora_generacion, :tipo_recaudador, :recaudador, :comprobante_pago_interno,
         :comprobante_pago_recaudador, :tipo_servicio, :tipo_documento, :numero_documento,
         :fecha_generacion, :fecha_operacion, :fecha_transaccion, :monto_recaudado,
         :monto_cuenta_1, :monto_cuenta_2)
    RETURNING id
    """
)


def _row(registro, fecha_transaccion, **overrides):
    base = {
        "registro": registro,
        "hora_generacion": time(10, 30, 0),
        "tipo_recaudador": "BAP",
        "recaudador": "BANCO DEL PACIFICO",
        "comprobante_pago_interno": f"INT-{registro}",
        "comprobante_pago_recaudador": f"REC-{registro}",
        "tipo_servicio": "CDP",
        "tipo_documento": "CON",
        "numero_documento": f"DOC-{registro}",
        "fecha_generacion": fecha_transaccion.date(),
        "fecha_operacion": fecha_transaccion,
        "fecha_transaccion": fecha_transaccion,
        "monto_recaudado": Decimal("119.00"),
        "monto_cuenta_1": Decimal("119.00"),
        "monto_cuenta_2": Decimal("0.00"),
    }
    base.update(overrides)
    return base


async def _seed_pagos(db_session, rows):
    ids = []
    for row in rows:
        result = await db_session.execute(INSERT_SQL, row)
        ids.append(result.scalar_one())
    await db_session.commit()
    return ids


@pytest_asyncio.fixture(autouse=True)
async def _cleanup_pagos(db_session):
    yield
    await db_session.execute(text("DELETE FROM axis.axis_pagos WHERE registro LIKE 'TEST-PAG-%'"))
    await db_session.commit()


@pytest.mark.asyncio
async def test_list_returns_items_within_range(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_pagos(
        db_session,
        [
            _row("TEST-PAG-101", datetime(2031, 6, 5, 9, 0, 0)),
            _row("TEST-PAG-102", datetime(2031, 6, 15, 9, 0, 0)),
            _row("TEST-PAG-103", datetime(2031, 6, 25, 9, 0, 0)),
        ],
    )

    response = await client.get(
        "/api/reportes/pagos",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 3
    assert body["page"] == 1
    assert body["page_size"] == 50
    registros = [item["registro"] for item in body["items"]]
    assert registros == ["TEST-PAG-103", "TEST-PAG-102", "TEST-PAG-101"]
    first = body["items"][0]
    assert first["recaudador"] == "BANCO DEL PACIFICO"
    assert first["monto_recaudado"] == 119.0
    assert first["tipo_documento_catalogo_item_id"] is None
    assert first["tipo_recaudador_catalogo_item_id"] is None
    assert first["tipo_servicio_catalogo_item_id"] is None


@pytest.mark.asyncio
async def test_list_allows_range_crossing_month(client, db_session):
    headers = await _auth_headers(client, db_session)

    response = await client.get(
        "/api/reportes/pagos",
        params={"fecha_desde": "2031-06-15", "fecha_hasta": "2031-07-05"},
        headers=headers,
    )

    assert response.status_code == 200


@pytest.mark.asyncio
async def test_list_rejects_desde_after_hasta(client, db_session):
    headers = await _auth_headers(client, db_session)

    response = await client.get(
        "/api/reportes/pagos",
        params={"fecha_desde": "2031-06-20", "fecha_hasta": "2031-06-10"},
        headers=headers,
    )

    assert response.status_code == 400


@pytest.mark.asyncio
async def test_list_excludes_soft_deleted_rows(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_pagos(
        db_session,
        [
            _row("TEST-PAG-201", datetime(2031, 6, 5, 9, 0, 0)),
            _row("TEST-PAG-202", datetime(2031, 6, 6, 9, 0, 0)),
        ],
    )
    await db_session.execute(
        text("UPDATE axis.axis_pagos SET deleted_at = now() WHERE registro = 'TEST-PAG-202'")
    )
    await db_session.commit()

    response = await client.get(
        "/api/reportes/pagos",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 1
    assert body["items"][0]["registro"] == "TEST-PAG-201"


@pytest.mark.asyncio
async def test_list_includes_rows_at_the_end_of_the_last_day(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_pagos(
        db_session,
        [
            _row("TEST-PAG-601", datetime(2031, 6, 30, 23, 59, 59)),
            _row("TEST-PAG-602", datetime(2031, 7, 1, 0, 0, 0)),
        ],
    )

    response = await client.get(
        "/api/reportes/pagos",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    registros = [item["registro"] for item in body["items"]]
    assert "TEST-PAG-601" in registros
    assert "TEST-PAG-602" not in registros


@pytest.mark.asyncio
async def test_list_pagination_page_two_offset(client, db_session):
    headers = await _auth_headers(client, db_session)
    base = datetime(2031, 5, 1, 8, 0, 0)
    rows = [_row(f"TEST-PAG-p-{i:03d}", base.replace(day=1 + i % 28)) for i in range(55)]
    await _seed_pagos(db_session, rows)

    response = await client.get(
        "/api/reportes/pagos",
        params={"fecha_desde": "2031-05-01", "fecha_hasta": "2031-06-25", "page": 2},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 55
    assert body["page"] == 2
    assert len(body["items"]) == 5


@pytest.mark.asyncio
async def test_list_out_of_range_page_returns_empty(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_pagos(db_session, [_row("TEST-PAG-301", datetime(2031, 6, 5, 9, 0, 0))])

    response = await client.get(
        "/api/reportes/pagos",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30", "page": 5},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["items"] == []
    assert body["total"] == 1


@pytest.mark.asyncio
async def test_list_without_token_returns_401(client, db_session):
    response = await client.get(
        "/api/reportes/pagos",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
    )
    assert response.status_code == 401


@pytest.mark.asyncio
async def test_list_blocked_when_must_change_password_is_true(client, db_session):
    user = User(
        email="pendiente@example.com",
        password_hash=hash_password("Sup3rSecret!"),
        full_name="Usuario Pendiente",
        must_change_password=True,
    )
    db_session.add(user)
    await db_session.commit()
    await db_session.refresh(user)
    token = create_access_token(user_id=user.id, email=user.email)

    response = await client.get(
        "/api/reportes/pagos",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 403
    assert response.json()["detail"]["code"] == "password_change_required"


EXPECTED_HEADERS = [
    "Registro",
    "Hora de Generación",
    "Tipo de Recaudador",
    "Recaudador",
    "Comprobante de Pago Interno",
    "Comprobante de Pago del Recaudador",
    "Tipo de Servicio",
    "Tipo de Documento",
    "Número de Documento",
    "Fecha de Generación",
    "Fecha de Operación",
    "Fecha de Transacción",
    "Monto Recaudado",
    "Monto Cuenta 1",
    "Monto Cuenta 2",
    "ID de Catálogo (Tipo de Documento)",
    "ID de Catálogo (Tipo de Recaudador)",
    "ID de Catálogo (Tipo de Servicio)",
]


@pytest.mark.asyncio
async def test_export_csv_returns_all_matching_rows(client, db_session):
    headers = await _auth_headers(client, db_session)
    rows = [
        _row(f"TEST-PAG-e-{i:03d}", datetime(2031, 8, 1 + i, 9, 0, 0))
        for i in range(28)
    ]
    await _seed_pagos(db_session, rows)

    response = await client.get(
        "/api/reportes/pagos/export",
        params={"fecha_desde": "2031-08-01", "fecha_hasta": "2031-08-31", "formato": "csv"},
        headers=headers,
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    assert "pagos_2031-08-01_2031-08-31.csv" in response.headers["content-disposition"]

    text_content = response.content.decode("utf-8-sig")
    lines = [line for line in text_content.splitlines() if line]
    reader = csv.reader(lines)
    parsed_rows = list(reader)
    assert parsed_rows[0] == EXPECTED_HEADERS
    assert len(parsed_rows[0]) == 18
    assert len(lines) - 1 == 28

    data_row = parsed_rows[1]
    assert len(data_row) == 18
    assert data_row[0].startswith("TEST-PAG-e-")


@pytest.mark.asyncio
async def test_export_xlsx_returns_all_matching_rows(client, db_session):
    headers = await _auth_headers(client, db_session)
    rows = [
        _row(f"TEST-PAG-x-{i:03d}", datetime(2031, 9, 1 + i, 9, 0, 0))
        for i in range(29)
    ]
    await _seed_pagos(db_session, rows)

    response = await client.get(
        "/api/reportes/pagos/export",
        params={"fecha_desde": "2031-09-01", "fecha_hasta": "2031-09-30", "formato": "xlsx"},
        headers=headers,
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "pagos_2031-09-01_2031-09-30.xlsx" in response.headers["content-disposition"]

    workbook = load_workbook(io.BytesIO(response.content))
    sheet = workbook.active
    header_row = [sheet.cell(row=1, column=col).value for col in range(1, 19)]
    assert header_row == EXPECTED_HEADERS
    assert sheet.max_row - 1 == 29

    data_row = [sheet.cell(row=2, column=col).value for col in range(1, 19)]
    assert data_row[0].startswith("TEST-PAG-x-")


@pytest.mark.asyncio
async def test_export_excludes_soft_deleted_rows(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_pagos(
        db_session,
        [
            _row("TEST-PAG-501", datetime(2031, 6, 5, 9, 0, 0)),
            _row("TEST-PAG-502", datetime(2031, 6, 6, 9, 0, 0)),
        ],
    )
    await db_session.execute(
        text("UPDATE axis.axis_pagos SET deleted_at = now() WHERE registro = 'TEST-PAG-502'")
    )
    await db_session.commit()

    response = await client.get(
        "/api/reportes/pagos/export",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30", "formato": "csv"},
        headers=headers,
    )

    assert response.status_code == 200
    text_content = response.content.decode("utf-8-sig")
    lines = [line for line in text_content.splitlines() if line]
    assert len(lines) - 1 == 1


@pytest.mark.asyncio
async def test_export_without_token_returns_401(client, db_session):
    response = await client.get(
        "/api/reportes/pagos/export",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30", "formato": "csv"},
    )
    assert response.status_code == 401


@pytest.mark.asyncio
async def test_list_truncates_datetime_columns_to_date_only(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_pagos(db_session, [_row("TEST-PAG-TRUNC-001", datetime(2031, 6, 5, 14, 35, 0))])

    response = await client.get(
        "/api/reportes/pagos",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
        headers=headers,
    )

    assert response.status_code == 200
    first = response.json()["items"][0]
    assert first["fecha_operacion"] == "2031-06-05"
    assert first["fecha_transaccion"] == "2031-06-05"


@pytest.mark.asyncio
async def test_export_truncates_datetime_columns_to_date_only(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_pagos(db_session, [_row("TEST-PAG-TRUNC-002", datetime(2031, 6, 6, 14, 35, 0))])

    response = await client.get(
        "/api/reportes/pagos/export",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30", "formato": "csv"},
        headers=headers,
    )

    assert response.status_code == 200
    text_content = response.content.decode("utf-8-sig")
    lines = [line for line in text_content.splitlines() if line]
    reader = csv.reader(lines)
    parsed_rows = list(reader)
    data_row = parsed_rows[-1]
    assert data_row[10] == "2031-06-06"
    assert data_row[11] == "2031-06-06"


async def _fecha_minima_actual(client, headers, path):
    response = await client.get(path, headers=headers)
    fecha_str = response.json()["fecha_minima"]
    if fecha_str is None:
        return datetime(1901, 1, 1)
    return datetime.strptime(fecha_str, "%Y-%m-%d") - timedelta(days=1)


@pytest.mark.asyncio
async def test_fecha_minima_includes_seeded_row(client, db_session):
    headers = await _auth_headers(client, db_session)
    fecha_anterior = await _fecha_minima_actual(client, headers, "/api/reportes/pagos/fecha-minima")
    await _seed_pagos(db_session, [_row("TEST-PAG-FMIN-001", fecha_anterior)])

    response = await client.get("/api/reportes/pagos/fecha-minima", headers=headers)

    assert response.status_code == 200
    assert response.json()["fecha_minima"] == fecha_anterior.strftime("%Y-%m-%d")


@pytest.mark.asyncio
async def test_fecha_minima_ignores_soft_deleted_rows(client, db_session):
    headers = await _auth_headers(client, db_session)
    fecha_anterior = await _fecha_minima_actual(client, headers, "/api/reportes/pagos/fecha-minima")
    await _seed_pagos(db_session, [_row("TEST-PAG-FMIN-101", fecha_anterior)])
    await db_session.execute(
        text("UPDATE axis.axis_pagos SET deleted_at = now() WHERE registro = 'TEST-PAG-FMIN-101'")
    )
    await db_session.commit()

    response = await client.get("/api/reportes/pagos/fecha-minima", headers=headers)

    assert response.status_code == 200
    assert response.json()["fecha_minima"] != fecha_anterior.strftime("%Y-%m-%d")


@pytest.mark.asyncio
async def test_fecha_minima_without_token_returns_401(client, db_session):
    response = await client.get("/api/reportes/pagos/fecha-minima")
    assert response.status_code == 401
