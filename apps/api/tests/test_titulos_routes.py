import csv
import io
from datetime import date, time, timedelta

import pytest
import pytest_asyncio
from decimal import Decimal
from openpyxl import load_workbook
from sqlalchemy import text

from app.auth import create_access_token, hash_password
from app.models import User


async def _create_user(db_session, email="user@example.com", password="Sup3rSecret!"):
    user = User(email=email, password_hash=hash_password(password), full_name="Test User")
    db_session.add(user)
    await db_session.commit()
    await db_session.refresh(user)
    return user


async def _auth_headers(client, db_session):
    await _create_user(db_session)
    response = await client.post(
        "/api/auth/login", json={"email": "user@example.com", "password": "Sup3rSecret!"}
    )
    token = response.json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


INSERT_PERSONA_SQL = text(
    """
    INSERT INTO axis.personas (identificacion, tipo_identificacion, nombre)
    VALUES (:identificacion, :tipo_identificacion, :nombre)
    ON CONFLICT (identificacion) DO NOTHING
    """
)

INSERT_SQL = text(
    """
    INSERT INTO axis.axis_titulos
        (registro, hora_generacion, codigo_titulo_credito, tipo_identificacion, identificacion,
         nombre_completo, etapa_cobranza, estado, codigo_referencia, concepto,
         nombre_elabora_titulo, nombre_solicita, nombre_aprobacion, motivo_anulacion,
         fecha_generacion, fecha_registro, fecha_elaboracion, fecha_solicitud, fecha_aprobacion,
         fecha_notificacion, fecha_pago, fecha_anulacion, valor, multas, interes, valor_total,
         estado_catalogo_item_id, etapa_cobranza_catalogo_item_id, tipo_identificacion_catalogo_item_id)
    VALUES
        (:registro, :hora_generacion, :codigo_titulo_credito, :tipo_identificacion, :identificacion,
         :nombre_completo, :etapa_cobranza, :estado, :codigo_referencia, :concepto,
         :nombre_elabora_titulo, :nombre_solicita, :nombre_aprobacion, :motivo_anulacion,
         :fecha_generacion, :fecha_registro, :fecha_elaboracion, :fecha_solicitud, :fecha_aprobacion,
         :fecha_notificacion, :fecha_pago, :fecha_anulacion, :valor, :multas, :interes, :valor_total,
         :estado_catalogo_item_id, :etapa_cobranza_catalogo_item_id, :tipo_identificacion_catalogo_item_id)
    RETURNING id
    """
)


def _row(registro, fecha_registro, identificacion="TEST-TIT-CED-0001", **overrides):
    base = {
        "registro": registro,
        "hora_generacion": time(9, 15, 0),
        "codigo_titulo_credito": f"TC-{registro}",
        "tipo_identificacion": "CED",
        "identificacion": identificacion,
        "nombre_completo": "Deudor de Prueba",
        "etapa_cobranza": "NOTIFICACION",
        "estado": "ACTIVO",
        "codigo_referencia": "REF-001",
        "concepto": "Concepto de prueba",
        "nombre_elabora_titulo": "Elaborador de Prueba",
        "nombre_solicita": "Solicitante de Prueba",
        "nombre_aprobacion": "Aprobador de Prueba",
        "motivo_anulacion": None,
        "fecha_generacion": fecha_registro,
        "fecha_registro": fecha_registro,
        "fecha_elaboracion": fecha_registro,
        "fecha_solicitud": fecha_registro,
        "fecha_aprobacion": None,
        "fecha_notificacion": None,
        "fecha_pago": None,
        "fecha_anulacion": None,
        "valor": Decimal("150.00"),
        "multas": Decimal("10.00"),
        "interes": Decimal("2.50"),
        "valor_total": Decimal("162.50"),
        "estado_catalogo_item_id": 12,
        "etapa_cobranza_catalogo_item_id": 5,
        "tipo_identificacion_catalogo_item_id": 3,
    }
    base.update(overrides)
    return base


async def _seed_personas(db_session, identificaciones):
    for ident in identificaciones:
        await db_session.execute(
            INSERT_PERSONA_SQL,
            {"identificacion": ident, "tipo_identificacion": "CED", "nombre": f"Persona {ident}"},
        )
    await db_session.commit()


async def _seed_titulos(db_session, rows):
    personas = {row["identificacion"] for row in rows}
    await _seed_personas(db_session, personas)
    ids = []
    for row in rows:
        result = await db_session.execute(INSERT_SQL, row)
        ids.append(result.scalar_one())
    await db_session.commit()
    return ids


@pytest_asyncio.fixture(autouse=True)
async def _cleanup_titulos(db_session):
    yield
    await db_session.execute(text("DELETE FROM axis.axis_titulos WHERE registro LIKE 'TEST-TIT-%'"))
    await db_session.execute(text("DELETE FROM axis.personas WHERE identificacion LIKE 'TEST-TIT-%'"))
    await db_session.commit()


@pytest.mark.asyncio
async def test_list_returns_items_within_range(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_titulos(
        db_session,
        [
            _row("TEST-TIT-101", date(2031, 6, 5), identificacion="TEST-TIT-CED-0001"),
            _row("TEST-TIT-102", date(2031, 6, 15), identificacion="TEST-TIT-CED-0002"),
            _row("TEST-TIT-103", date(2031, 6, 25), identificacion="TEST-TIT-CED-0003"),
        ],
    )

    response = await client.get(
        "/api/reportes/titulos",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 3
    assert body["page"] == 1
    assert body["page_size"] == 50
    registros = [item["registro"] for item in body["items"]]
    assert registros == ["TEST-TIT-103", "TEST-TIT-102", "TEST-TIT-101"]
    first = body["items"][0]
    assert first["nombre_completo"] == "Deudor de Prueba"
    assert first["valor_total"] == 162.5
    assert first["estado_catalogo_item_id"] == 12
    assert first["fecha_registro"] == "2031-06-25"


@pytest.mark.asyncio
async def test_list_allows_range_crossing_month(client, db_session):
    headers = await _auth_headers(client, db_session)

    response = await client.get(
        "/api/reportes/titulos",
        params={"fecha_desde": "2031-06-15", "fecha_hasta": "2031-07-05"},
        headers=headers,
    )

    assert response.status_code == 200


@pytest.mark.asyncio
async def test_list_rejects_desde_after_hasta(client, db_session):
    headers = await _auth_headers(client, db_session)

    response = await client.get(
        "/api/reportes/titulos",
        params={"fecha_desde": "2031-06-20", "fecha_hasta": "2031-06-10"},
        headers=headers,
    )

    assert response.status_code == 400


@pytest.mark.asyncio
async def test_list_excludes_soft_deleted_rows(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_titulos(
        db_session,
        [
            _row("TEST-TIT-201", date(2031, 6, 5), identificacion="TEST-TIT-CED-0004"),
            _row("TEST-TIT-202", date(2031, 6, 6), identificacion="TEST-TIT-CED-0005"),
        ],
    )
    await db_session.execute(
        text("UPDATE axis.axis_titulos SET deleted_at = now() WHERE registro = 'TEST-TIT-202'")
    )
    await db_session.commit()

    response = await client.get(
        "/api/reportes/titulos",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 1
    assert body["items"][0]["registro"] == "TEST-TIT-201"


@pytest.mark.asyncio
async def test_list_pagination_page_two_offset(client, db_session):
    headers = await _auth_headers(client, db_session)
    base = date(2031, 5, 1)
    rows = [
        _row(f"TEST-TIT-p-{i:03d}", base + timedelta(days=i), identificacion=f"TEST-TIT-CED-p{i:03d}")
        for i in range(55)
    ]
    await _seed_titulos(db_session, rows)

    response = await client.get(
        "/api/reportes/titulos",
        params={"fecha_desde": "2031-05-01", "fecha_hasta": "2031-06-25", "page": 2},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 55
    assert body["page"] == 2
    assert len(body["items"]) == 5
    assert body["items"][0]["registro"] == "TEST-TIT-p-004"
    assert body["items"][-1]["registro"] == "TEST-TIT-p-000"


@pytest.mark.asyncio
async def test_list_out_of_range_page_returns_empty(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_titulos(db_session, [_row("TEST-TIT-301", date(2031, 6, 5))])

    response = await client.get(
        "/api/reportes/titulos",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30", "page": 5},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["items"] == []
    assert body["total"] == 1


@pytest.mark.asyncio
async def test_list_without_token_returns_401(client, db_session):
    response = await client.get(
        "/api/reportes/titulos",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
    )
    assert response.status_code == 401


@pytest.mark.asyncio
async def test_list_blocked_when_must_change_password_is_true(client, db_session):
    user = User(
        email="pendiente@example.com",
        password_hash=hash_password("Sup3rSecret!"),
        full_name="Usuario Pendiente",
        must_change_password=True,
    )
    db_session.add(user)
    await db_session.commit()
    await db_session.refresh(user)
    token = create_access_token(user_id=user.id, email=user.email)

    response = await client.get(
        "/api/reportes/titulos",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 403
    assert response.json()["detail"]["code"] == "password_change_required"


EXPECTED_HEADERS = [
    "Registro",
    "Hora de Generación del Registro",
    "Código Título Crédito",
    "Tipo de Identificación",
    "Identificación",
    "Nombre Completo",
    "Etapa Cobranza",
    "Estado",
    "Código de Referencia",
    "Concepto",
    "Nombre Elabora Título de Crédito",
    "Nombre que Solicita",
    "Nombre de Aprobación",
    "Motivo de Anulación",
    "Fecha de Generación del Registro",
    "Fecha de Registro",
    "Fecha de Elaboración",
    "Fecha de Solicitud",
    "Fecha de Aprobación",
    "Fecha de Notificación",
    "Fecha de Pago",
    "Fecha de Anulación",
    "Valor",
    "Multas",
    "Interés",
    "Valor Total",
    "ID de Catálogo (Estado)",
    "ID de Catálogo (Etapa de Cobranza)",
    "ID de Catálogo (Tipo de Identificación)",
]


@pytest.mark.asyncio
async def test_export_csv_returns_all_matching_rows(client, db_session):
    headers = await _auth_headers(client, db_session)
    base = date(2031, 8, 1)
    rows = [
        _row(f"TEST-TIT-e-{i:03d}", base + timedelta(days=i), identificacion=f"TEST-TIT-CED-e{i:03d}")
        for i in range(31)
    ]
    await _seed_titulos(db_session, rows)

    response = await client.get(
        "/api/reportes/titulos/export",
        params={"fecha_desde": "2031-08-01", "fecha_hasta": "2031-08-31", "formato": "csv"},
        headers=headers,
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    assert "titulos_2031-08-01_2031-08-31.csv" in response.headers["content-disposition"]

    text_content = response.content.decode("utf-8-sig")
    lines = [line for line in text_content.splitlines() if line]
    reader = csv.reader(lines)
    parsed_rows = list(reader)
    assert parsed_rows[0] == EXPECTED_HEADERS
    assert len(parsed_rows[0]) == 29
    assert len(lines) - 1 == 31

    data_row = parsed_rows[1]
    assert len(data_row) == 29
    assert data_row[0].startswith("TEST-TIT-e-")


@pytest.mark.asyncio
async def test_export_xlsx_returns_all_matching_rows(client, db_session):
    headers = await _auth_headers(client, db_session)
    base = date(2031, 9, 1)
    rows = [
        _row(f"TEST-TIT-x-{i:03d}", base + timedelta(days=i), identificacion=f"TEST-TIT-CED-x{i:03d}")
        for i in range(30)
    ]
    await _seed_titulos(db_session, rows)

    response = await client.get(
        "/api/reportes/titulos/export",
        params={"fecha_desde": "2031-09-01", "fecha_hasta": "2031-09-30", "formato": "xlsx"},
        headers=headers,
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "titulos_2031-09-01_2031-09-30.xlsx" in response.headers["content-disposition"]

    workbook = load_workbook(io.BytesIO(response.content))
    sheet = workbook.active
    header_row = [sheet.cell(row=1, column=col).value for col in range(1, 30)]
    assert header_row == EXPECTED_HEADERS
    assert sheet.max_row - 1 == 30

    data_row = [sheet.cell(row=2, column=col).value for col in range(1, 30)]
    assert data_row[0].startswith("TEST-TIT-x-")


@pytest.mark.asyncio
async def test_export_excludes_soft_deleted_rows(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_titulos(
        db_session,
        [
            _row("TEST-TIT-501", date(2031, 6, 5), identificacion="TEST-TIT-CED-0006"),
            _row("TEST-TIT-502", date(2031, 6, 6), identificacion="TEST-TIT-CED-0007"),
        ],
    )
    await db_session.execute(
        text("UPDATE axis.axis_titulos SET deleted_at = now() WHERE registro = 'TEST-TIT-502'")
    )
    await db_session.commit()

    response = await client.get(
        "/api/reportes/titulos/export",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30", "formato": "csv"},
        headers=headers,
    )

    assert response.status_code == 200
    text_content = response.content.decode("utf-8-sig")
    lines = [line for line in text_content.splitlines() if line]
    assert len(lines) - 1 == 1


@pytest.mark.asyncio
async def test_export_without_token_returns_401(client, db_session):
    response = await client.get(
        "/api/reportes/titulos/export",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30", "formato": "csv"},
    )
    assert response.status_code == 401
