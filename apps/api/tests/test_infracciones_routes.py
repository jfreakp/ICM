import csv
import io
from datetime import datetime, timedelta
from decimal import Decimal

import pytest
import pytest_asyncio
from openpyxl import load_workbook
from sqlalchemy import text

from app.auth import create_access_token, hash_password
from app.models import User


async def _create_user(db_session, email="user@example.com", password="Sup3rSecret!"):
    user = User(email=email, password_hash=hash_password(password), full_name="Test User")
    db_session.add(user)
    await db_session.commit()
    await db_session.refresh(user)
    return user


async def _auth_headers(client, db_session):
    await _create_user(db_session)
    response = await client.post(
        "/api/auth/login", json={"email": "user@example.com", "password": "Sup3rSecret!"}
    )
    token = response.json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


INSERT_PERSONA_SQL = text(
    """
    INSERT INTO axis.personas (identificacion, tipo_identificacion, nombre)
    VALUES (:identificacion, :tipo_identificacion, :nombre)
    ON CONFLICT (identificacion) DO NOTHING
    """
)

INSERT_SQL = text(
    """
    INSERT INTO axis.axis_infracciones
        (registro, fecha_registro, fecha_emision, fecha_aprobacion, fecha_vencimiento, estado,
         codigo_infraccion, codigo_infraccion_ant, contravencion, articulo, literal,
         descripcion_articulo, periodo_fiscal, oficina, origen_registro, tipo_registro_infraccion,
         tipo_emision, tipo_deudor, codigo_usuario_registra, observacion, provincia, localidad,
         lugar_infraccion, canal, placa, tipo_identificacion_infractor,
         numero_identificacion_infractor, nombre_infractor, tipo_identificacion_propietario,
         numero_identificacion_propietario, nombre_propietario, indicador_bloqueada,
         indicador_acta_juzgamiento, indicador_modificada, indicador_calcula_recargo,
         valor_capital, valor_capital_exonerado, valor_recargo, valor_recargo_exonerado,
         valor_intereses, valor_total)
    VALUES
        (:registro, :fecha_registro, :fecha_emision, :fecha_aprobacion, :fecha_vencimiento, :estado,
         :codigo_infraccion, :codigo_infraccion_ant, :contravencion, :articulo, :literal,
         :descripcion_articulo, :periodo_fiscal, :oficina, :origen_registro, :tipo_registro_infraccion,
         :tipo_emision, :tipo_deudor, :codigo_usuario_registra, :observacion, :provincia, :localidad,
         :lugar_infraccion, :canal, :placa, :tipo_identificacion_infractor,
         :numero_identificacion_infractor, :nombre_infractor, :tipo_identificacion_propietario,
         :numero_identificacion_propietario, :nombre_propietario, :indicador_bloqueada,
         :indicador_acta_juzgamiento, :indicador_modificada, :indicador_calcula_recargo,
         :valor_capital, :valor_capital_exonerado, :valor_recargo, :valor_recargo_exonerado,
         :valor_intereses, :valor_total)
    RETURNING id
    """
)


def _row(registro, fecha_registro, estado="EMITIDA", **overrides):
    base = {
        "registro": registro,
        "fecha_registro": fecha_registro,
        "fecha_emision": fecha_registro,
        "fecha_aprobacion": fecha_registro,
        "fecha_vencimiento": fecha_registro,
        "estado": estado,
        "codigo_infraccion": f"COD-{registro}",
        "codigo_infraccion_ant": f"ANT-{registro}",
        "contravencion": f"CONTRA-{registro}",
        "articulo": "139",
        "literal": "1",
        "descripcion_articulo": "Descripción de prueba",
        "periodo_fiscal": "ACTUAL",
        "oficina": "GAD LOJA",
        "origen_registro": "AXIS",
        "tipo_registro_infraccion": "PARTE",
        "tipo_emision": "ACT",
        "tipo_deudor": "CONDUCTOR",
        "codigo_usuario_registra": "USR001",
        "observacion": "Observación de prueba",
        "provincia": "LOJ",
        "localidad": "LOJ",
        "lugar_infraccion": "Av. de prueba",
        "canal": "APP",
        "placa": "ABC1234",
        "tipo_identificacion_infractor": "CED",
        "numero_identificacion_infractor": "TEST-INF-CED-0001",
        "nombre_infractor": "Infractor de Prueba",
        "tipo_identificacion_propietario": "CED",
        "numero_identificacion_propietario": "TEST-INF-CED-0001",
        "nombre_propietario": "Propietario de Prueba",
        "indicador_bloqueada": "N",
        "indicador_acta_juzgamiento": "N",
        "indicador_modificada": "N",
        "indicador_calcula_recargo": "S",
        "valor_capital": Decimal("50.00"),
        "valor_capital_exonerado": Decimal("0.00"),
        "valor_recargo": Decimal("5.00"),
        "valor_recargo_exonerado": Decimal("0.00"),
        "valor_intereses": Decimal("1.00"),
        "valor_total": Decimal("56.00"),
    }
    base.update(overrides)
    return base


async def _seed_personas(db_session, identificaciones):
    """Insert test personas that infracciones will reference"""
    for ident in identificaciones:
        await db_session.execute(
            INSERT_PERSONA_SQL,
            {"identificacion": ident, "tipo_identificacion": "CED", "nombre": f"Persona {ident}"}
        )
    await db_session.commit()


async def _seed_infracciones(db_session, rows):
    # Ensure personas exist first
    personas = set()
    for row in rows:
        personas.add(row.get("numero_identificacion_infractor"))
        personas.add(row.get("numero_identificacion_propietario"))
    personas.discard(None)
    if personas:
        await _seed_personas(db_session, personas)

    ids = []
    for row in rows:
        result = await db_session.execute(INSERT_SQL, row)
        ids.append(result.scalar_one())
    await db_session.commit()
    return ids


@pytest_asyncio.fixture(autouse=True)
async def _cleanup_infracciones(db_session):
    yield
    await db_session.execute(text("DELETE FROM axis.axis_infracciones WHERE registro LIKE 'TEST-INF-%'"))
    await db_session.execute(text("DELETE FROM axis.personas WHERE identificacion LIKE 'TEST-INF-%'"))
    await db_session.commit()


@pytest.mark.asyncio
async def test_estados_returns_distinct_values(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_infracciones(
        db_session,
        [
            _row("TEST-INF-001", datetime(2024, 3, 5), estado="EMITIDA"),
            _row("TEST-INF-002", datetime(2024, 3, 6), estado="PAGADA"),
            _row("TEST-INF-003", datetime(2024, 3, 7), estado="EMITIDA"),
        ],
    )

    response = await client.get("/api/reportes/infracciones/estados", headers=headers)

    assert response.status_code == 200
    body = response.json()
    assert "EMITIDA" in body
    assert "PAGADA" in body


@pytest.mark.asyncio
async def test_estados_without_token_returns_401(client, db_session):
    response = await client.get("/api/reportes/infracciones/estados")
    assert response.status_code == 401


@pytest.mark.asyncio
async def test_list_returns_items_within_range(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_infracciones(
        db_session,
        [
            _row("TEST-INF-101", datetime(2031, 6, 5), estado="EMITIDA", observacion="Primera"),
            _row("TEST-INF-102", datetime(2031, 6, 15), estado="EMITIDA", observacion="Segunda"),
            _row("TEST-INF-103", datetime(2031, 6, 25), estado="EMITIDA", observacion="Tercera"),
        ],
    )

    response = await client.get(
        "/api/reportes/infracciones",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 3
    assert body["page"] == 1
    assert body["page_size"] == 50
    registros = [item["registro"] for item in body["items"]]
    assert registros == ["TEST-INF-103", "TEST-INF-102", "TEST-INF-101"]
    first = body["items"][0]
    assert first["estado"] == "EMITIDA"
    assert first["observacion"] == "Tercera"
    assert first["valor_total"] == 56.0


@pytest.mark.asyncio
async def test_list_allows_range_crossing_month(client, db_session):
    headers = await _auth_headers(client, db_session)

    response = await client.get(
        "/api/reportes/infracciones",
        params={"fecha_desde": "2024-06-15", "fecha_hasta": "2024-07-05"},
        headers=headers,
    )

    assert response.status_code == 200


@pytest.mark.asyncio
async def test_list_rejects_desde_after_hasta(client, db_session):
    headers = await _auth_headers(client, db_session)

    response = await client.get(
        "/api/reportes/infracciones",
        params={"fecha_desde": "2024-06-20", "fecha_hasta": "2024-06-10"},
        headers=headers,
    )

    assert response.status_code == 400


@pytest.mark.asyncio
async def test_list_filters_by_estado(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_infracciones(
        db_session,
        [
            _row("TEST-INF-201", datetime(2031, 6, 5), estado="EMITIDA"),
            _row("TEST-INF-202", datetime(2031, 6, 6), estado="PAGADA"),
        ],
    )

    response = await client.get(
        "/api/reportes/infracciones",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30", "estado": "PAGADA"},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 1
    assert body["items"][0]["registro"] == "TEST-INF-202"


@pytest.mark.asyncio
async def test_list_pagination_page_two_offset(client, db_session):
    headers = await _auth_headers(client, db_session)
    base = datetime(2031, 5, 1)
    rows = [
        _row(f"TEST-INF-p-{i:03d}", base + timedelta(minutes=30 * i), estado="EMITIDA")
        for i in range(55)
    ]
    await _seed_infracciones(db_session, rows)

    response = await client.get(
        "/api/reportes/infracciones",
        params={"fecha_desde": "2031-05-01", "fecha_hasta": "2031-05-31", "page": 2},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 55
    assert body["page"] == 2
    assert len(body["items"]) == 5
    assert body["items"][0]["registro"] == "TEST-INF-p-004"
    assert body["items"][-1]["registro"] == "TEST-INF-p-000"


@pytest.mark.asyncio
async def test_list_out_of_range_page_returns_empty(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_infracciones(db_session, [_row("TEST-INF-301", datetime(2031, 6, 5))])

    response = await client.get(
        "/api/reportes/infracciones",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30", "page": 5},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["items"] == []
    assert body["total"] == 1


@pytest.mark.asyncio
async def test_list_without_token_returns_401(client, db_session):
    response = await client.get(
        "/api/reportes/infracciones",
        params={"fecha_desde": "2024-06-01", "fecha_hasta": "2024-06-30"},
    )
    assert response.status_code == 401


EXPECTED_HEADERS = [
    "Registro",
    "Fecha de Registro",
    "Fecha de Emisión",
    "Fecha de Aprobación",
    "Fecha de Vencimiento",
    "Estado",
    "Código de Infracción",
    "Código de Infracción Anterior",
    "Contravención",
    "Artículo",
    "Literal",
    "Descripción del Artículo",
    "Período Fiscal",
    "Oficina",
    "Origen de Registro",
    "Tipo de Registro",
    "Tipo de Emisión",
    "Tipo de Deudor",
    "Usuario que Registra",
    "Observación",
    "Provincia",
    "Localidad",
    "Lugar de Infracción",
    "Canal",
    "Placa",
    "Tipo de Identificación (Infractor)",
    "Número de Identificación (Infractor)",
    "Nombre del Infractor",
    "Tipo de Identificación (Propietario)",
    "Número de Identificación (Propietario)",
    "Nombre del Propietario",
    "Bloqueada",
    "Acta de Juzgamiento",
    "Modificada",
    "Calcula Recargo",
    "Valor Capital",
    "Valor Capital Exonerado",
    "Valor Recargo",
    "Valor Recargo Exonerado",
    "Valor Intereses",
    "Valor Total",
    "Hora de Generación del Registro",
    "Fecha de Generación del Registro",
    "Tipo de Infracción",
    "Código del Usuario que Aprueba",
    "Código del Usuario que Notifica",
    "Tipo de Licencia",
    "Zona",
    "Distrito",
    "Circuito",
    "Dispositivo",
    "Geo-referencia-X",
    "Geo-referencia-Y",
    "Tipo de Identificación del Agente",
    "Número de Identificación del Agente",
    "Nombre del Agente",
    "Código del Agente de Tránsito",
    "Tipo de Infracción (2)",
    "Código de la Infracción Origen",
    "Código de la Empresa del Convenio",
    "Porcentaje Principal",
    "Porcentaje Convenio",
    "Cuenta Bancaria Principal",
    "Cuenta Bancaria Convenio",
    "Fecha de Notificación",
    "Fecha de Pago",
    "Fecha de Impugnación",
    "Fecha de Convenio",
    "Fecha de Anulación",
    "Fecha de Coactiva",
    "ID de Catálogo (Canal)",
    "ID de Catálogo (Estado)",
    "ID de Catálogo (Localidad)",
    "ID de Catálogo (Origen de Registro)",
    "ID de Catálogo (Provincia)",
    "ID de Catálogo (Tipo de Deudor)",
    "ID de Catálogo (Tipo de Emisión)",
    "ID de Catálogo (Tipo de Identificación del Agente)",
    "ID de Catálogo (Tipo de Identificación del Infractor)",
    "ID de Catálogo (Tipo de Identificación del Propietario)",
    "ID de Catálogo (Tipo de Licencia)",
    "ID de Catálogo (Tipo de Registro de Infracción)",
    "ID de Catálogo (Zona)",
]


@pytest.mark.asyncio
async def test_export_csv_returns_all_matching_rows(client, db_session):
    headers = await _auth_headers(client, db_session)
    base = datetime(2031, 8, 1)
    rows = [
        _row(f"TEST-INF-e-{i:03d}", base + timedelta(minutes=30 * i), estado="EMITIDA")
        for i in range(55)
    ]
    await _seed_infracciones(db_session, rows)

    response = await client.get(
        "/api/reportes/infracciones/export",
        params={"fecha_desde": "2031-08-01", "fecha_hasta": "2031-08-31", "formato": "csv"},
        headers=headers,
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    assert "infracciones_2031-08-01_2031-08-31.csv" in response.headers["content-disposition"]

    text_content = response.content.decode("utf-8-sig")
    lines = [line for line in text_content.splitlines() if line]
    reader = csv.reader(lines)
    parsed_rows = list(reader)
    assert parsed_rows[0] == EXPECTED_HEADERS
    assert len(parsed_rows[0]) == 83
    assert len(lines) - 1 == 55

    data_row = parsed_rows[1]
    assert len(data_row) == 83
    assert data_row[5] == "EMITIDA"
    assert data_row[0].startswith("TEST-INF-e-")


@pytest.mark.asyncio
async def test_export_xlsx_returns_all_matching_rows(client, db_session):
    headers = await _auth_headers(client, db_session)
    base = datetime(2031, 9, 1)
    rows = [
        _row(f"TEST-INF-x-{i:03d}", base + timedelta(minutes=30 * i), estado="EMITIDA")
        for i in range(55)
    ]
    await _seed_infracciones(db_session, rows)

    response = await client.get(
        "/api/reportes/infracciones/export",
        params={"fecha_desde": "2031-09-01", "fecha_hasta": "2031-09-30", "formato": "xlsx"},
        headers=headers,
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "infracciones_2031-09-01_2031-09-30.xlsx" in response.headers["content-disposition"]

    workbook = load_workbook(io.BytesIO(response.content))
    sheet = workbook.active
    header_row = [sheet.cell(row=1, column=col).value for col in range(1, 84)]
    assert header_row == EXPECTED_HEADERS
    assert sheet.max_row - 1 == 55

    data_row = [sheet.cell(row=2, column=col).value for col in range(1, 84)]
    assert data_row[5] == "EMITIDA"
    assert data_row[0].startswith("TEST-INF-x-")


@pytest.mark.asyncio
async def test_export_without_token_returns_401(client, db_session):
    response = await client.get(
        "/api/reportes/infracciones/export",
        params={"fecha_desde": "2024-06-01", "fecha_hasta": "2024-06-30", "formato": "csv"},
    )
    assert response.status_code == 401


@pytest.mark.asyncio
async def test_list_estados_infracciones_blocked_when_must_change_password_is_true(client, db_session):
    from app.auth import create_access_token, hash_password
    from app.models import User

    user = User(
        email="pendiente@example.com",
        password_hash=hash_password("Sup3rSecret!"),
        full_name="Usuario Pendiente",
        must_change_password=True,
    )
    db_session.add(user)
    await db_session.commit()
    await db_session.refresh(user)
    token = create_access_token(user_id=user.id, email=user.email)

    response = await client.get(
        "/api/reportes/infracciones/estados", headers={"Authorization": f"Bearer {token}"}
    )

    assert response.status_code == 403
    assert response.json()["detail"]["code"] == "password_change_required"


@pytest.mark.asyncio
async def test_list_truncates_datetime_columns_to_date_only(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_infracciones(
        db_session, [_row("TEST-INF-TRUNC-001", datetime(2031, 6, 5, 14, 35, 0), estado="EMITIDA")]
    )

    response = await client.get(
        "/api/reportes/infracciones",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
        headers=headers,
    )

    assert response.status_code == 200
    first = response.json()["items"][0]
    assert first["fecha_registro"] == "2031-06-05"
    assert first["fecha_emision"] == "2031-06-05"
    assert first["fecha_aprobacion"] == "2031-06-05"
    assert first["fecha_vencimiento"] == "2031-06-05"


@pytest.mark.asyncio
async def test_export_truncates_datetime_columns_to_date_only(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_infracciones(
        db_session, [_row("TEST-INF-TRUNC-002", datetime(2031, 6, 6, 14, 35, 0), estado="EMITIDA")]
    )

    response = await client.get(
        "/api/reportes/infracciones/export",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30", "formato": "csv"},
        headers=headers,
    )

    assert response.status_code == 200
    text_content = response.content.decode("utf-8-sig")
    lines = [line for line in text_content.splitlines() if line]
    reader = csv.reader(lines)
    parsed_rows = list(reader)
    data_row = parsed_rows[-1]
    assert data_row[1] == "2031-06-06"
    assert data_row[2] == "2031-06-06"
    assert data_row[3] == "2031-06-06"
    assert data_row[4] == "2031-06-06"


@pytest.mark.asyncio
async def test_list_shows_newly_added_columns(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_infracciones(
        db_session, [_row("TEST-INF-NEWCOL-001", datetime(2031, 6, 5, 14, 35, 0), estado="EMITIDA")]
    )
    await _seed_personas(db_session, ["TEST-INF-AGT-0001"])
    await db_session.execute(
        text(
            """
            UPDATE axis.axis_infracciones
            SET hora_generacion = '14:35:00',
                fecha_generacion = '2031-06-01',
                tipo_infraccion = 'DE TRANSITO',
                zona = 'ZONA 1',
                tipo_identificacion_agente = 'CED',
                numero_identificacion_agente = 'TEST-INF-AGT-0001',
                nombre_agente = 'Agente de Prueba',
                fecha_coactiva = :fecha_coactiva,
                estado_catalogo_item_id = 77
            WHERE registro = 'TEST-INF-NEWCOL-001'
            """
        ),
        {"fecha_coactiva": datetime(2031, 6, 12, 8, 0, 0)},
    )
    await db_session.commit()

    response = await client.get(
        "/api/reportes/infracciones",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
        headers=headers,
    )

    assert response.status_code == 200
    first = response.json()["items"][0]
    assert first["hora_generacion"] == "14:35:00"
    assert first["fecha_generacion"] == "2031-06-01"
    assert first["tipo_infraccion"] == "DE TRANSITO"
    assert first["zona"] == "ZONA 1"
    assert first["numero_identificacion_agente"] == "TEST-INF-AGT-0001"
    assert first["nombre_agente"] == "Agente de Prueba"
    assert first["fecha_coactiva"] == "2031-06-12"
    assert first["estado_catalogo_item_id"] == 77


async def _fecha_minima_actual(client, headers, path):
    response = await client.get(path, headers=headers)
    fecha_str = response.json()["fecha_minima"]
    if fecha_str is None:
        return datetime(1901, 1, 1)
    return datetime.strptime(fecha_str, "%Y-%m-%d") - timedelta(days=1)


@pytest.mark.asyncio
async def test_fecha_minima_includes_seeded_row(client, db_session):
    headers = await _auth_headers(client, db_session)
    fecha_anterior = await _fecha_minima_actual(client, headers, "/api/reportes/infracciones/fecha-minima")
    await _seed_infracciones(db_session, [_row("TEST-INF-FMIN-001", fecha_anterior, estado="EMITIDA")])

    response = await client.get("/api/reportes/infracciones/fecha-minima", headers=headers)

    assert response.status_code == 200
    assert response.json()["fecha_minima"] == fecha_anterior.strftime("%Y-%m-%d")


@pytest.mark.asyncio
async def test_fecha_minima_ignores_soft_deleted_rows(client, db_session):
    headers = await _auth_headers(client, db_session)
    fecha_anterior = await _fecha_minima_actual(client, headers, "/api/reportes/infracciones/fecha-minima")
    await _seed_infracciones(db_session, [_row("TEST-INF-FMIN-101", fecha_anterior, estado="EMITIDA")])
    await db_session.execute(
        text("UPDATE axis.axis_infracciones SET deleted_at = now() WHERE registro = 'TEST-INF-FMIN-101'")
    )
    await db_session.commit()

    response = await client.get("/api/reportes/infracciones/fecha-minima", headers=headers)

    assert response.status_code == 200
    assert response.json()["fecha_minima"] != fecha_anterior.strftime("%Y-%m-%d")


@pytest.mark.asyncio
async def test_fecha_minima_without_token_returns_401(client, db_session):
    response = await client.get("/api/reportes/infracciones/fecha-minima")
    assert response.status_code == 401
