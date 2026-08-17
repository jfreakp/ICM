import csv
import io
from datetime import date, datetime, time

import pytest
import pytest_asyncio
from openpyxl import load_workbook
from sqlalchemy import text

from app.auth import create_access_token, hash_password
from app.models import User


async def _create_user(db_session, email="user@example.com", password="Sup3rSecret!"):
    user = User(email=email, password_hash=hash_password(password), full_name="Test User")
    db_session.add(user)
    await db_session.commit()
    await db_session.refresh(user)
    return user


async def _auth_headers(client, db_session):
    await _create_user(db_session)
    response = await client.post(
        "/api/auth/login", json={"email": "user@example.com", "password": "Sup3rSecret!"}
    )
    token = response.json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


INSERT_PERSONA_SQL = text(
    """
    INSERT INTO axis.personas (identificacion, tipo_identificacion, nombre)
    VALUES (:identificacion, :tipo_identificacion, :nombre)
    ON CONFLICT (identificacion) DO NOTHING
    """
)

INSERT_SQL = text(
    """
    INSERT INTO axis.axis_crv
        (registro, hora_generacion, codigo_orden_crv, codigo_actividad, codigo_oficina,
         descripcion_oficina, placa, nombre_agente, identificacion_agente, motivo_ingreso_crv,
         clase, provincia, localidad_ciudad, ciudadela, area, direccion, remolque, km_remolque,
         valor_remolque, fecha_generacion, fecha_ingreso, fecha_salida,
         localidad_ciudad_catalogo_item_id, provincia_catalogo_item_id)
    VALUES
        (:registro, :hora_generacion, :codigo_orden_crv, :codigo_actividad, :codigo_oficina,
         :descripcion_oficina, :placa, :nombre_agente, :identificacion_agente, :motivo_ingreso_crv,
         :clase, :provincia, :localidad_ciudad, :ciudadela, :area, :direccion, :remolque, :km_remolque,
         :valor_remolque, :fecha_generacion, :fecha_ingreso, :fecha_salida,
         :localidad_ciudad_catalogo_item_id, :provincia_catalogo_item_id)
    RETURNING id
    """
)


def _row(registro, fecha_ingreso, identificacion_agente="TEST-CRV-CED-0001", **overrides):
    base = {
        "registro": registro,
        "hora_generacion": time(8, 45, 0),
        "codigo_orden_crv": f"ORD-{registro}",
        "codigo_actividad": "ING",
        "codigo_oficina": "OF-01",
        "descripcion_oficina": "Oficina Centro",
        "placa": "ABC1234",
        "nombre_agente": "Agente de Prueba",
        "identificacion_agente": identificacion_agente,
        "motivo_ingreso_crv": "INFRACCION",
        "clase": "LIVIANO",
        "provincia": "LOJ",
        "localidad_ciudad": "LOJA",
        "ciudadela": "CENTRO",
        "area": "URBANA",
        "direccion": "Av. de Prueba",
        "remolque": "GRUA-01",
        "km_remolque": "5.2",
        "valor_remolque": "15.00",
        "fecha_generacion": fecha_ingreso.date(),
        "fecha_ingreso": fecha_ingreso,
        "fecha_salida": None,
        "localidad_ciudad_catalogo_item_id": 8,
        "provincia_catalogo_item_id": 4,
    }
    base.update(overrides)
    return base


async def _seed_personas(db_session, identificaciones):
    for ident in identificaciones:
        await db_session.execute(
            INSERT_PERSONA_SQL,
            {"identificacion": ident, "tipo_identificacion": "CED", "nombre": f"Persona {ident}"},
        )
    await db_session.commit()


async def _seed_crv(db_session, rows):
    personas = {row["identificacion_agente"] for row in rows}
    await _seed_personas(db_session, personas)
    ids = []
    for row in rows:
        result = await db_session.execute(INSERT_SQL, row)
        ids.append(result.scalar_one())
    await db_session.commit()
    return ids


@pytest_asyncio.fixture(autouse=True)
async def _cleanup_crv(db_session):
    yield
    await db_session.execute(text("DELETE FROM axis.axis_crv WHERE registro LIKE 'TEST-CRV-%'"))
    await db_session.execute(text("DELETE FROM axis.personas WHERE identificacion LIKE 'TEST-CRV-%'"))
    await db_session.commit()


@pytest.mark.asyncio
async def test_list_returns_items_within_range(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_crv(
        db_session,
        [
            _row("TEST-CRV-101", datetime(2031, 6, 5, 9, 0, 0), identificacion_agente="TEST-CRV-CED-0001"),
            _row("TEST-CRV-102", datetime(2031, 6, 15, 9, 0, 0), identificacion_agente="TEST-CRV-CED-0002"),
            _row("TEST-CRV-103", datetime(2031, 6, 25, 9, 0, 0), identificacion_agente="TEST-CRV-CED-0003"),
        ],
    )

    response = await client.get(
        "/api/reportes/crv",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 3
    assert body["page"] == 1
    assert body["page_size"] == 50
    registros = [item["registro"] for item in body["items"]]
    assert registros == ["TEST-CRV-103", "TEST-CRV-102", "TEST-CRV-101"]
    first = body["items"][0]
    assert first["nombre_agente"] == "Agente de Prueba"
    assert first["identificacion_agente"] == "TEST-CRV-CED-0003"
    assert first["localidad_ciudad_catalogo_item_id"] == 8
    assert first["provincia_catalogo_item_id"] == 4


@pytest.mark.asyncio
async def test_list_allows_range_crossing_month(client, db_session):
    headers = await _auth_headers(client, db_session)

    response = await client.get(
        "/api/reportes/crv",
        params={"fecha_desde": "2031-06-15", "fecha_hasta": "2031-07-05"},
        headers=headers,
    )

    assert response.status_code == 200


@pytest.mark.asyncio
async def test_list_rejects_desde_after_hasta(client, db_session):
    headers = await _auth_headers(client, db_session)

    response = await client.get(
        "/api/reportes/crv",
        params={"fecha_desde": "2031-06-20", "fecha_hasta": "2031-06-10"},
        headers=headers,
    )

    assert response.status_code == 400


@pytest.mark.asyncio
async def test_list_excludes_soft_deleted_rows(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_crv(
        db_session,
        [
            _row("TEST-CRV-201", datetime(2031, 6, 5, 9, 0, 0), identificacion_agente="TEST-CRV-CED-0004"),
            _row("TEST-CRV-202", datetime(2031, 6, 6, 9, 0, 0), identificacion_agente="TEST-CRV-CED-0005"),
        ],
    )
    await db_session.execute(
        text("UPDATE axis.axis_crv SET deleted_at = now() WHERE registro = 'TEST-CRV-202'")
    )
    await db_session.commit()

    response = await client.get(
        "/api/reportes/crv",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 1
    assert body["items"][0]["registro"] == "TEST-CRV-201"


@pytest.mark.asyncio
async def test_list_includes_rows_at_the_end_of_the_last_day(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_crv(
        db_session,
        [
            _row("TEST-CRV-601", datetime(2031, 6, 30, 23, 59, 59), identificacion_agente="TEST-CRV-CED-0006"),
            _row("TEST-CRV-602", datetime(2031, 7, 1, 0, 0, 0), identificacion_agente="TEST-CRV-CED-0007"),
        ],
    )

    response = await client.get(
        "/api/reportes/crv",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    registros = [item["registro"] for item in body["items"]]
    assert "TEST-CRV-601" in registros
    assert "TEST-CRV-602" not in registros


@pytest.mark.asyncio
async def test_list_pagination_page_two_offset(client, db_session):
    headers = await _auth_headers(client, db_session)
    base = datetime(2031, 5, 1, 8, 0, 0)
    rows = [_row(f"TEST-CRV-p-{i:03d}", base.replace(day=1 + i % 28)) for i in range(55)]
    await _seed_crv(db_session, rows)

    response = await client.get(
        "/api/reportes/crv",
        params={"fecha_desde": "2031-05-01", "fecha_hasta": "2031-06-25", "page": 2},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 55
    assert body["page"] == 2
    assert len(body["items"]) == 5


@pytest.mark.asyncio
async def test_list_out_of_range_page_returns_empty(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_crv(db_session, [_row("TEST-CRV-301", datetime(2031, 6, 5, 9, 0, 0))])

    response = await client.get(
        "/api/reportes/crv",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30", "page": 5},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["items"] == []
    assert body["total"] == 1


@pytest.mark.asyncio
async def test_list_without_token_returns_401(client, db_session):
    response = await client.get(
        "/api/reportes/crv",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
    )
    assert response.status_code == 401


@pytest.mark.asyncio
async def test_list_blocked_when_must_change_password_is_true(client, db_session):
    user = User(
        email="pendiente@example.com",
        password_hash=hash_password("Sup3rSecret!"),
        full_name="Usuario Pendiente",
        must_change_password=True,
    )
    db_session.add(user)
    await db_session.commit()
    await db_session.refresh(user)
    token = create_access_token(user_id=user.id, email=user.email)

    response = await client.get(
        "/api/reportes/crv",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
        headers={"Authorization": f"Bearer {token}"},
    )

    assert response.status_code == 403
    assert response.json()["detail"]["code"] == "password_change_required"


EXPECTED_HEADERS = [
    "Registro",
    "Hora de Generación del Registro",
    "Código de Orden CRV",
    "Código de Actividad",
    "Código de Oficina",
    "Descripción de Oficina",
    "Placa",
    "Nombre Agente",
    "Identificación de Agente",
    "Motivo Ingreso al CRV",
    "Clase",
    "Provincia",
    "Localidad o Ciudad",
    "Ciudadela",
    "Área",
    "Dirección",
    "Remolque",
    "Km de Remolque",
    "Valor Remolque",
    "Fecha de Generación del Registro",
    "Fecha Ingreso",
    "Fecha Salida",
    "ID de Catálogo (Localidad o Ciudad)",
    "ID de Catálogo (Provincia)",
]


@pytest.mark.asyncio
async def test_export_csv_returns_all_matching_rows(client, db_session):
    headers = await _auth_headers(client, db_session)
    rows = [
        _row(f"TEST-CRV-e-{i:03d}", datetime(2031, 8, 1 + i, 9, 0, 0))
        for i in range(28)
    ]
    await _seed_crv(db_session, rows)

    response = await client.get(
        "/api/reportes/crv/export",
        params={"fecha_desde": "2031-08-01", "fecha_hasta": "2031-08-31", "formato": "csv"},
        headers=headers,
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    assert "crv_2031-08-01_2031-08-31.csv" in response.headers["content-disposition"]

    text_content = response.content.decode("utf-8-sig")
    lines = [line for line in text_content.splitlines() if line]
    reader = csv.reader(lines)
    parsed_rows = list(reader)
    assert parsed_rows[0] == EXPECTED_HEADERS
    assert len(parsed_rows[0]) == 24
    assert len(lines) - 1 == 28

    data_row = parsed_rows[1]
    assert len(data_row) == 24
    assert data_row[0].startswith("TEST-CRV-e-")


@pytest.mark.asyncio
async def test_export_xlsx_returns_all_matching_rows(client, db_session):
    headers = await _auth_headers(client, db_session)
    rows = [
        _row(f"TEST-CRV-x-{i:03d}", datetime(2031, 9, 1 + i, 9, 0, 0))
        for i in range(29)
    ]
    await _seed_crv(db_session, rows)

    response = await client.get(
        "/api/reportes/crv/export",
        params={"fecha_desde": "2031-09-01", "fecha_hasta": "2031-09-30", "formato": "xlsx"},
        headers=headers,
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "crv_2031-09-01_2031-09-30.xlsx" in response.headers["content-disposition"]

    workbook = load_workbook(io.BytesIO(response.content))
    sheet = workbook.active
    header_row = [sheet.cell(row=1, column=col).value for col in range(1, 25)]
    assert header_row == EXPECTED_HEADERS
    assert sheet.max_row - 1 == 29

    data_row = [sheet.cell(row=2, column=col).value for col in range(1, 25)]
    assert data_row[0].startswith("TEST-CRV-x-")


@pytest.mark.asyncio
async def test_export_excludes_soft_deleted_rows(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_crv(
        db_session,
        [
            _row("TEST-CRV-501", datetime(2031, 6, 5, 9, 0, 0), identificacion_agente="TEST-CRV-CED-0008"),
            _row("TEST-CRV-502", datetime(2031, 6, 6, 9, 0, 0), identificacion_agente="TEST-CRV-CED-0009"),
        ],
    )
    await db_session.execute(
        text("UPDATE axis.axis_crv SET deleted_at = now() WHERE registro = 'TEST-CRV-502'")
    )
    await db_session.commit()

    response = await client.get(
        "/api/reportes/crv/export",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30", "formato": "csv"},
        headers=headers,
    )

    assert response.status_code == 200
    text_content = response.content.decode("utf-8-sig")
    lines = [line for line in text_content.splitlines() if line]
    assert len(lines) - 1 == 1


@pytest.mark.asyncio
async def test_export_without_token_returns_401(client, db_session):
    response = await client.get(
        "/api/reportes/crv/export",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30", "formato": "csv"},
    )
    assert response.status_code == 401


@pytest.mark.asyncio
async def test_list_truncates_datetime_columns_to_date_only(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_crv(
        db_session,
        [
            _row(
                "TEST-CRV-TRUNC-001",
                datetime(2031, 6, 5, 14, 35, 0),
                fecha_salida=datetime(2031, 6, 7, 10, 0, 0),
            )
        ],
    )

    response = await client.get(
        "/api/reportes/crv",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
        headers=headers,
    )

    assert response.status_code == 200
    first = response.json()["items"][0]
    assert first["fecha_generacion"] == "2031-06-05"
    assert first["fecha_ingreso"] == "2031-06-05"
    assert first["fecha_salida"] == "2031-06-07"


@pytest.mark.asyncio
async def test_export_truncates_datetime_columns_to_date_only(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_crv(
        db_session,
        [
            _row(
                "TEST-CRV-TRUNC-002",
                datetime(2031, 6, 6, 14, 35, 0),
                fecha_salida=datetime(2031, 6, 8, 9, 0, 0),
            )
        ],
    )

    response = await client.get(
        "/api/reportes/crv/export",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30", "formato": "csv"},
        headers=headers,
    )

    assert response.status_code == 200
    text_content = response.content.decode("utf-8-sig")
    lines = [line for line in text_content.splitlines() if line]
    reader = csv.reader(lines)
    parsed_rows = list(reader)
    data_row = parsed_rows[-1]
    assert data_row[19] == "2031-06-06"
    assert data_row[20] == "2031-06-06"
    assert data_row[21] == "2031-06-08"
