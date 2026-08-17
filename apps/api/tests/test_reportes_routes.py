from datetime import datetime
from decimal import Decimal

import pytest
import pytest_asyncio
from sqlalchemy import text

from app.auth import create_access_token, hash_password
from app.models import User


async def _create_user(db_session, email="user@example.com", password="Sup3rSecret!"):
    user = User(email=email, password_hash=hash_password(password), full_name="Test User")
    db_session.add(user)
    await db_session.commit()
    await db_session.refresh(user)
    return user


async def _auth_headers(client, db_session):
    await _create_user(db_session)
    response = await client.post(
        "/api/auth/login", json={"email": "user@example.com", "password": "Sup3rSecret!"}
    )
    token = response.json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


INSERT_SQL = text(
    """
    INSERT INTO axis.axis_impugnaciones
        (registro, fecha_registro, fecha_acta, estado, codigo_infraccion_axis,
         contravencion, tipo_acta, articulo_original, monto_capital_original, observacion)
    VALUES
        (:registro, :fecha_registro, :fecha_acta, :estado, :codigo_infraccion_axis,
         :contravencion, :tipo_acta, :articulo_original, :monto_capital_original, :observacion)
    RETURNING id
    """
)


async def _seed_impugnaciones(db_session, rows):
    ids = []
    for row in rows:
        result = await db_session.execute(INSERT_SQL, row)
        ids.append(result.scalar_one())
    await db_session.commit()
    return ids


@pytest_asyncio.fixture(autouse=True)
async def _cleanup_impugnaciones(db_session):
    yield
    await db_session.execute(text("DELETE FROM axis.axis_impugnaciones WHERE registro LIKE 'TEST-%'"))
    await db_session.commit()


def _row(registro, fecha_registro, estado="A", **overrides):
    base = {
        "registro": registro,
        "fecha_registro": fecha_registro,
        "fecha_acta": fecha_registro,
        "estado": estado,
        "codigo_infraccion_axis": "COD-1",
        "contravencion": "Contravencion de prueba",
        "tipo_acta": "Tipo A",
        "articulo_original": "Art 1",
        "monto_capital_original": Decimal("100.00"),
        "observacion": "Observación de prueba",
    }
    base.update(overrides)
    return base


@pytest.mark.asyncio
async def test_estados_returns_distinct_values(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_impugnaciones(
        db_session,
        [
            _row("TEST-001", datetime(2024, 3, 5), estado="A"),
            _row("TEST-002", datetime(2024, 3, 6), estado="B"),
            _row("TEST-003", datetime(2024, 3, 7), estado="A"),
        ],
    )

    response = await client.get("/api/reportes/impugnaciones/estados", headers=headers)

    assert response.status_code == 200
    body = response.json()
    assert "A" in body
    assert "B" in body


@pytest.mark.asyncio
async def test_estados_without_token_returns_401(client, db_session):
    response = await client.get("/api/reportes/impugnaciones/estados")
    assert response.status_code == 401


from datetime import timedelta


@pytest.mark.asyncio
async def test_list_returns_items_within_range(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_impugnaciones(
        db_session,
        [
            _row("TEST-101", datetime(2031, 6, 5), estado="A", observacion="Primera"),
            _row("TEST-102", datetime(2031, 6, 15), estado="A", observacion="Segunda"),
            _row("TEST-103", datetime(2031, 6, 25), estado="A", observacion="Tercera"),
        ],
    )

    response = await client.get(
        "/api/reportes/impugnaciones",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 3
    assert body["page"] == 1
    assert body["page_size"] == 50
    registros = [item["registro"] for item in body["items"]]
    assert registros == ["TEST-103", "TEST-102", "TEST-101"]
    first = body["items"][0]
    assert first["estado"] == "A"
    assert first["monto_capital_original"] == 100.0
    assert first["observacion"] == "Tercera"


@pytest.mark.asyncio
async def test_list_allows_range_crossing_month(client, db_session):
    headers = await _auth_headers(client, db_session)

    response = await client.get(
        "/api/reportes/impugnaciones",
        params={"fecha_desde": "2024-06-15", "fecha_hasta": "2024-07-05"},
        headers=headers,
    )

    assert response.status_code == 200


@pytest.mark.asyncio
async def test_list_rejects_desde_after_hasta(client, db_session):
    headers = await _auth_headers(client, db_session)

    response = await client.get(
        "/api/reportes/impugnaciones",
        params={"fecha_desde": "2024-06-20", "fecha_hasta": "2024-06-10"},
        headers=headers,
    )

    assert response.status_code == 400


@pytest.mark.asyncio
async def test_list_filters_by_estado(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_impugnaciones(
        db_session,
        [
            _row("TEST-201", datetime(2031, 6, 5), estado="A"),
            _row("TEST-202", datetime(2031, 6, 6), estado="B"),
        ],
    )

    response = await client.get(
        "/api/reportes/impugnaciones",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30", "estado": "B"},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 1
    assert body["items"][0]["registro"] == "TEST-202"


@pytest.mark.asyncio
async def test_list_pagination_page_two_offset(client, db_session):
    headers = await _auth_headers(client, db_session)
    base = datetime(2031, 5, 1)
    rows = [
        _row(f"TEST-p-{i:03d}", base + timedelta(minutes=30 * i), estado="A")
        for i in range(55)
    ]
    await _seed_impugnaciones(db_session, rows)

    response = await client.get(
        "/api/reportes/impugnaciones",
        params={"fecha_desde": "2031-05-01", "fecha_hasta": "2031-05-31", "page": 2},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 55
    assert body["page"] == 2
    assert len(body["items"]) == 5
    assert body["items"][0]["registro"] == "TEST-p-004"
    assert body["items"][-1]["registro"] == "TEST-p-000"


@pytest.mark.asyncio
async def test_list_out_of_range_page_returns_empty(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_impugnaciones(db_session, [_row("TEST-301", datetime(2031, 6, 5))])

    response = await client.get(
        "/api/reportes/impugnaciones",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30", "page": 5},
        headers=headers,
    )

    assert response.status_code == 200
    body = response.json()
    assert body["items"] == []
    assert body["total"] == 1


@pytest.mark.asyncio
async def test_list_without_token_returns_401(client, db_session):
    response = await client.get(
        "/api/reportes/impugnaciones",
        params={"fecha_desde": "2024-06-01", "fecha_hasta": "2024-06-30"},
    )
    assert response.status_code == 401


import csv
import io

from openpyxl import load_workbook

EXPECTED_HEADERS = [
    "Registro",
    "Fecha de Registro",
    "Fecha de Acta",
    "Estado",
    "Código de Infracción AXIS",
    "Contravención",
    "Tipo de Acta",
    "Artículo Original",
    "Monto Capital Original",
    "Observación",
    "Hora de Generación del Registro",
    "Fecha de Generación del Registro",
    "Número de Crédito",
    "Número de Trámite",
    "Código de la Infracción Generada en AXIS Cloud",
    "Juzgado",
    "Código de la Provincia",
    "Código de la Localidad",
    "Número del Proceso",
    "Monto Modificado por la Sentencia",
    "Puntos Original",
    "Puntos Modificados por la Sentencia",
    "Literal Original",
    "Artículo Modificado por la Sentencia",
    "Literal Modificado por la Sentencia",
    "Fecha de Vencimiento Original",
    "Fecha de Vencimiento Modificado por la Sentencia",
    "Sanción Original",
    "Sanción Modificada por la Sentencia",
    "Código del Usuario",
    "Código del Usuario que Aprueba",
    "Número de Acta de Juzgamiento",
    "Fecha de Aprobación",
    "Fecha de Anulación",
    "Código de Usuario que Anula",
    "Observación de Anulación",
    "ID de Catálogo (Artículo Original)",
    "ID de Catálogo (Artículo Modificado por la Sentencia)",
    "ID de Catálogo (Localidad)",
    "ID de Catálogo (Provincia)",
    "ID de Catálogo (Tipo de Acta)",
]


@pytest.mark.asyncio
async def test_export_csv_returns_all_matching_rows(client, db_session):
    headers = await _auth_headers(client, db_session)
    base = datetime(2031, 8, 1)
    rows = [
        _row(f"TEST-e-{i:03d}", base + timedelta(minutes=30 * i), estado="A")
        for i in range(55)
    ]
    await _seed_impugnaciones(db_session, rows)

    response = await client.get(
        "/api/reportes/impugnaciones/export",
        params={"fecha_desde": "2031-08-01", "fecha_hasta": "2031-08-31", "formato": "csv"},
        headers=headers,
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    assert "impugnaciones_2031-08-01_2031-08-31.csv" in response.headers["content-disposition"]

    text = response.content.decode("utf-8-sig")
    lines = [line for line in text.splitlines() if line]
    reader = csv.reader(lines)
    parsed_rows = list(reader)
    assert parsed_rows[0] == EXPECTED_HEADERS
    assert len(lines) - 1 == 55

    data_row = parsed_rows[1]
    assert len(data_row) == 41
    assert data_row[3] == "A"
    assert data_row[0].startswith("TEST-e-")


@pytest.mark.asyncio
async def test_export_xlsx_returns_all_matching_rows(client, db_session):
    headers = await _auth_headers(client, db_session)
    base = datetime(2031, 9, 1)
    rows = [
        _row(f"TEST-x-{i:03d}", base + timedelta(minutes=30 * i), estado="A")
        for i in range(55)
    ]
    await _seed_impugnaciones(db_session, rows)

    response = await client.get(
        "/api/reportes/impugnaciones/export",
        params={"fecha_desde": "2031-09-01", "fecha_hasta": "2031-09-30", "formato": "xlsx"},
        headers=headers,
    )

    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "impugnaciones_2031-09-01_2031-09-30.xlsx" in response.headers["content-disposition"]

    workbook = load_workbook(io.BytesIO(response.content))
    sheet = workbook.active
    header_row = [sheet.cell(row=1, column=col).value for col in range(1, 42)]
    assert header_row == EXPECTED_HEADERS
    assert sheet.max_row - 1 == 55

    data_row = [sheet.cell(row=2, column=col).value for col in range(1, 42)]
    assert data_row[3] == "A"
    assert data_row[0].startswith("TEST-x-")


@pytest.mark.asyncio
async def test_export_without_token_returns_401(client, db_session):
    response = await client.get(
        "/api/reportes/impugnaciones/export",
        params={"fecha_desde": "2024-06-01", "fecha_hasta": "2024-06-30", "formato": "csv"},
    )
    assert response.status_code == 401


@pytest.mark.asyncio
async def test_list_truncates_datetime_columns_to_date_only(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_impugnaciones(
        db_session, [_row("TEST-TRUNC-001", datetime(2031, 6, 5, 14, 35, 0), estado="A")]
    )

    response = await client.get(
        "/api/reportes/impugnaciones",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
        headers=headers,
    )

    assert response.status_code == 200
    first = response.json()["items"][0]
    assert first["fecha_registro"] == "2031-06-05"
    assert first["fecha_acta"] == "2031-06-05"


@pytest.mark.asyncio
async def test_export_truncates_datetime_columns_to_date_only(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_impugnaciones(
        db_session, [_row("TEST-TRUNC-002", datetime(2031, 6, 6, 14, 35, 0), estado="A")]
    )

    response = await client.get(
        "/api/reportes/impugnaciones/export",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30", "formato": "csv"},
        headers=headers,
    )

    assert response.status_code == 200
    text_content = response.content.decode("utf-8-sig")
    lines = [line for line in text_content.splitlines() if line]
    reader = csv.reader(lines)
    parsed_rows = list(reader)
    data_row = parsed_rows[-1]
    assert data_row[1] == "2031-06-06"
    assert data_row[2] == "2031-06-06"


@pytest.mark.asyncio
async def test_list_shows_newly_added_columns(client, db_session):
    headers = await _auth_headers(client, db_session)
    await _seed_impugnaciones(
        db_session, [_row("TEST-NEWCOL-001", datetime(2031, 6, 5, 14, 35, 0), estado="A")]
    )
    await db_session.execute(
        text(
            """
            UPDATE axis.axis_impugnaciones
            SET hora_generacion = '14:35:00',
                fecha_generacion = '2031-06-01',
                numero_credito = 'CRED-001',
                juzgado = 'Juzgado de Prueba',
                fecha_anulacion = :fecha_anulacion,
                tipo_acta_catalogo_item_id = 42
            WHERE registro = 'TEST-NEWCOL-001'
            """
        ),
        {"fecha_anulacion": datetime(2031, 6, 10, 9, 0, 0)},
    )
    await db_session.commit()

    response = await client.get(
        "/api/reportes/impugnaciones",
        params={"fecha_desde": "2031-06-01", "fecha_hasta": "2031-06-30"},
        headers=headers,
    )

    assert response.status_code == 200
    first = response.json()["items"][0]
    assert first["hora_generacion"] == "14:35:00"
    assert first["fecha_generacion"] == "2031-06-01"
    assert first["numero_credito"] == "CRED-001"
    assert first["juzgado"] == "Juzgado de Prueba"
    assert first["fecha_anulacion"] == "2031-06-10"
    assert first["tipo_acta_catalogo_item_id"] == 42


@pytest.mark.asyncio
async def test_list_estados_blocked_when_must_change_password_is_true(client, db_session):
    from app.auth import create_access_token, hash_password
    from app.models import User

    user = User(
        email="pendiente@example.com",
        password_hash=hash_password("Sup3rSecret!"),
        full_name="Usuario Pendiente",
        must_change_password=True,
    )
    db_session.add(user)
    await db_session.commit()
    await db_session.refresh(user)
    token = create_access_token(user_id=user.id, email=user.email)

    response = await client.get(
        "/api/reportes/impugnaciones/estados", headers={"Authorization": f"Bearer {token}"}
    )

    assert response.status_code == 403
    assert response.json()["detail"]["code"] == "password_change_required"


async def _fecha_minima_actual(client, headers, path):
    response = await client.get(path, headers=headers)
    fecha_str = response.json()["fecha_minima"]
    if fecha_str is None:
        return datetime(1901, 1, 1)
    return datetime.strptime(fecha_str, "%Y-%m-%d") - timedelta(days=1)


@pytest.mark.asyncio
async def test_fecha_minima_includes_seeded_row(client, db_session):
    headers = await _auth_headers(client, db_session)
    fecha_anterior = await _fecha_minima_actual(client, headers, "/api/reportes/impugnaciones/fecha-minima")
    await _seed_impugnaciones(db_session, [_row("TEST-FMIN-001", fecha_anterior, estado="A")])

    response = await client.get("/api/reportes/impugnaciones/fecha-minima", headers=headers)

    assert response.status_code == 200
    assert response.json()["fecha_minima"] == fecha_anterior.strftime("%Y-%m-%d")


@pytest.mark.asyncio
async def test_fecha_minima_ignores_soft_deleted_rows(client, db_session):
    headers = await _auth_headers(client, db_session)
    fecha_anterior = await _fecha_minima_actual(client, headers, "/api/reportes/impugnaciones/fecha-minima")
    await _seed_impugnaciones(db_session, [_row("TEST-FMIN-101", fecha_anterior, estado="A")])
    await db_session.execute(
        text("UPDATE axis.axis_impugnaciones SET deleted_at = now() WHERE registro = 'TEST-FMIN-101'")
    )
    await db_session.commit()

    response = await client.get("/api/reportes/impugnaciones/fecha-minima", headers=headers)

    assert response.status_code == 200
    assert response.json()["fecha_minima"] != fecha_anterior.strftime("%Y-%m-%d")


@pytest.mark.asyncio
async def test_fecha_minima_without_token_returns_401(client, db_session):
    response = await client.get("/api/reportes/impugnaciones/fecha-minima")
    assert response.status_code == 401
